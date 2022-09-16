import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import pandas as pd
import re

# Import command line arguments (these can be automatically generated from the sample sheet using sample_sheet_reader.py)
parser = argparse.ArgumentParser(
    description="Parses meta data from provided excel file and runs SNP haplotyper"
)

# File input/output data
parser.add_argument(
    "-i",
    "--input_file",
    type=str,
    help="Excel file containing SNP Array meta data",
)

args = parser.parse_args()


def load_workbook_range(worksheet_and_range, ws):
    worksheet_string = worksheet_and_range[0]
    range_string = worksheet_and_range[1]

    # Remove absolute symbols from string
    range_string = range_string.replace("$", "")
    if len(re.findall("[A-Z]+", range_string)) == 2:
        col_start, col_end = re.findall("[A-Z]+", range_string)
        data_rows = []
        for row in ws[range_string]:
            data_rows.append([cell.value for cell in row])
        range_contents = pd.DataFrame(
            data_rows, columns=get_column_interval(col_start, col_end)
        )
    elif len(re.findall("[A-Z]+", range_string)) == 1:
        col_start = re.findall("[A-Z]+", range_string)
        col_end = col_start
        range_contents = ws[range_string].value
    return range_contents


wb = load_workbook(
    filename=args.input_file,
    read_only=True,
    keep_vba=False,
    data_only=True,
    keep_links=True,
)

argument_dict = {}

# Get list of defined ranges from provided excel sheet
defined_ranges = wb.defined_names
print(defined_ranges)
for dn in wb.defined_names.definedName:
    print(dn.name)
    print(dn.attr_text)
    if dn.attr_text is not None:
        for worksheet_and_range in wb.defined_names[dn.name].destinations:
            ws = wb.active
            range_values = load_workbook_range(worksheet_and_range, ws)
            argument_dict[dn.name] = range_values


biopsy_number = argument_dict["biopsy_number"]
chromosome = argument_dict["chromosome"]
consanguineous = argument_dict["consanguineous"]
de_novo = argument_dict["de_novo"]

# Flatten merged cells so only the entered value is returned
disease = argument_dict["disease"].values.tolist()[0][0]
disease_omim = argument_dict["disease_omim"].values.tolist()[0][0]

# Extract embryo data into dataframe dropping any empty rows
embryo_data_df = argument_dict["embryo_data"].dropna(axis=0)
embryo_data_df.columns = [
    "biopsy_no",
    "embryo_id",
    "embryo_sex",
    "embryo_column_name",
]

exclusion = argument_dict["exclusion"]
female_partner_hosp_num = argument_dict["female_partner_hosp_num"]
flanking_region_size = argument_dict["flanking_region_size"]
gene = argument_dict["gene"]
gene_end = argument_dict["gene_end"]
gene_omim = argument_dict["gene_omim"]
gene_start = argument_dict["gene_start"]
input_file = argument_dict["input_file"]
maternal_mutation = argument_dict["maternal_mutation"]
mode_of_inheritance = argument_dict["mode_of_inheritance"]
multi_analysis = argument_dict["multi_analysis"]
(
    partner1_type,
    partner1_sex,
    partner1_forename,
    partner1_surname,
    partner1_dob,
    partner1_DNA_number,
    partner1_column_name,
) = argument_dict["partner1_details"].values.tolist()[0]
(
    partner2_type,
    partner2_sex,
    partner2_forename,
    partner2_surname,
    partner2_dob,
    partner2_DNA_number,
    partner2_column_name,
) = argument_dict["partner2_details"].values.tolist()[0]
paste_gene = argument_dict["paste_gene"]
paternal_mutation = argument_dict["paternal_mutation"]
pgd_worksheet = argument_dict["pgd_worksheet"]
pgd_worksheet_denovo = argument_dict["pgd_worksheet_denovo"]
pru = argument_dict["pru"]
reference = argument_dict["reference"]
ref_relationship = argument_dict["ref_relationship"]
ref_relationship_to_couple = argument_dict["ref_relationship_to_couple"]
ref_seq = argument_dict["ref_seq"]
ref_status = argument_dict["ref_status"]
template_version = argument_dict["template_version"]

if partner1_sex=="male" and partner2_sex=="female":
    male_partner_status = partner1_type
    male_partner_col = partner1_column_name
    female_partner_status = partner2_type
    female_partner_col = partner2_column_name
elif partner1_sex=="female" and partner2_sex=="male":
    male_partner_status = partner2_type
    male_partner_col = partner2_column_name
    female_partner_status = partner1_type
    female_partner_col = partner1_column_name


python_location = "S:\Genetics_Data2\Array\Software\python-3.10.0-embed-amd64\python.exe"
snp_haplotype_script = "placeholder"
output_folder = "."

# Run SNP haplotyping script.
print(mode_of_inheritance)
if mode_of_inheritance == "Autosomal_Dominant":
    print(
        (f" {python_location} -i {snp_haplotype_script}"
        f" --input_file {input_file} --output_folder {output_folder}"
        f" --output_prefix {output_prefix} --mode_of_inheritance {mode_of_inheritance}"
        f" --male_partner {male_partner_col} --male_partner_status {male_partner_status}"
        f" --female_partner {female_partner_col} --female_partner_status {female_partner_status}"
        f" --reference {reference_col} --reference_status {reference_status}"
        f" --reference_relationship {reference_relationship} --embryo_ids 24.F4.EMB11.rhchp 25.F4.EMB12.rhchp 26.F4.EMB13.rhchp"
        f" --embryo_sex unknown unknown unknown --gene_symbol {gene_symbol} --gene_start {gene_start}"
        f" --gene_end {gene_end} --chr {chromosome}")
    )
elif mode_of_inheritance == "Autosomal_Recessive":
    print(
        (f" {python_location} -i {snp_haplotype_script}"
        f" --input_file {input_file} --output_folder {output_folder}"
        f" --output_prefix {output_prefix} --mode_of_inheritance {mode_of_inheritance}"
        f" --male_partner {male_partner_col} --male_partner_status {male_partner_status}"
        f" --female_partner {female_partner_col} --female_partner_status {female_partner_status}"
        f" --reference {reference_col} --reference_status {reference_status}"
        f" --reference_relationship {reference_relationship} --embryo_ids {embryo_ids}"
        f" --embryo_sex {embryo_sex} --gene_symbol {gene_symbol} --gene_start {gene_start}"
        f" --gene_end {gene_end} --chr {chromosome}")
    )
elif mode_of_inheritance == "x_linked":
    print(
        (f"/usr/bin/env /home/graeme/miniconda3/envs/python3.10env/bin/python"
        f"/home/graeme/.vscode/extensions/ms-python.python-2022.12.0/pythonFiles/lib/python/debugpy/adapter/../../debugpy/launcher 41117"
        f" -- src/snp_haplotyper/snp_haplotype.py --input_file test_data/autosomal_dominant/F4_BRCA2_AD.txt --output_folder output/ "
        f" --output_prefix F4_BRCA2_AD --mode_of_inheritance autosomal_dominant --male_partner 22.F4.MP.rhchp --male_partner_status affected "
        f"--female_partner 23.F4.FP.rhchp --female_partner_status unaffected --reference 19.F4.PGF.rhchp --reference_status affected "
        f"--reference_relationship grandparent --embryo_ids 24.F4.EMB11.rhchp 25.F4.EMB12.rhchp 26.F4.EMB13.rhchp"
        f" --embryo_sex unknown unknown unknown --gene_symbol BRCA2 --gene_start 32315086 --gene_end 32400268 --chr 13")
    )
else:
    pass  # raise exception