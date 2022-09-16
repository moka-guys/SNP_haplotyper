import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import pandas as pd
import re
import json


# Import command line arguments
parser = argparse.ArgumentParser(
    description="Populate excel spreadsheet template with test data from JSON file"
)

# File input/output data
parser.add_argument(
    "-t",
    "--template_file",
    type=str,
    help="Excel template file for entering SNP Array meta data",
)

parser.add_argument(
    "-j",
    "--json_file",
    type=str,
    help="JSON file containing SNP Array parameters for test data",
)

args = parser.parse_args()


def read_launch_json(json_file):
    # Import launch.json and parse the file for the correct parameters for each of the samples
    json_file_path = json_file

    with open(json_file_path, "r") as j:
        launch_contents = json.loads(j.read())

    # Grep for arguments
    p = re.compile(r"'name': '.*?,")
    run_names = p.findall(str(launch_contents["configurations"]))
    p = re.compile(r"'args': \[.*?\]")
    run_args = p.findall(str(launch_contents["configurations"]))

    run_data_dictionary = {}
    for i in range(len(run_args)):
        run_name = run_names[i].lstrip("'name : ").rstrip("', ")
        arg_list = (
            run_args[i].lstrip("'args': [").rstrip(']"').replace("'", "").split(", ")
        )
        arg_dictionary = {}
        dict_values = []
        for i in arg_list:
            if i.startswith("-"):
                dict_values = []
                dict_key = i.strip("-")
            else:
                dict_values.append(i)
            if len(dict_values) == 1:
                arg_dictionary[dict_key] = dict_values[0]  # Don't return a list of 1
            else:
                arg_dictionary[dict_key] = dict_values

        # Ensure that embryo_ids are a list even if only a single embryo is present
        if isinstance(arg_dictionary["embryo_ids"], str):
            embryo_ids_list = [arg_dictionary["embryo_ids"]]
        else:
            embryo_ids_list = arg_dictionary["embryo_ids"]

        if isinstance(arg_dictionary["embryo_sex"], str):
            embryo_sex_list = [arg_dictionary["embryo_sex"]]
        else:
            embryo_sex_list = arg_dictionary["embryo_sex"]

        run_data_dictionary[run_name] = args

        return run_data_dictionary


def load_workbook_range(worksheet_and_range, ws):
    worksheet_string = worksheet_and_range[0]
    range_string = worksheet_and_range[1]

    # Remove absolute symbols from string
    range_string = range_string.replace("$", "")
    if len(re.findall("[A-Z]+", range_string)) == 2:
        col_start, col_end = re.findall("[A-Z]+", range_string)
        data_rows = []
        for row in ws[range_string]:
            data_rows.append([cell.value for cell in row])
        range_contents = pd.DataFrame(
            data_rows, columns=get_column_interval(col_start, col_end)
        )
    elif len(re.findall("[A-Z]+", range_string)) == 1:
        col_start = re.findall("[A-Z]+", range_string)
        col_end = col_start
        range_contents = ws[range_string].value
    return range_contents


wb = load_workbook(
    filename=args.input_file,
    read_only=True,
    keep_vba=False,
    data_only=True,
    keep_links=True,
)

argument_dict = {}

# Get list of defined ranges from provided excel sheet
defined_ranges = wb.defined_names
print(defined_ranges)
for dn in wb.defined_names.definedName:
    print(dn.name)
    print(dn.attr_text)
    if dn.attr_text is not None:
        for worksheet_and_range in wb.defined_names[dn.name].destinations:
            ws = wb.active
            range_values = load_workbook_range(worksheet_and_range, ws)
            argument_dict[dn.name] = range_values


biopsy_number = argument_dict["biopsy_number"]
chromosome = argument_dict["chromosome"]
consanguineous = argument_dict["consanguineous"]
de_novo = argument_dict["de_novo"]

# Flatten merged cells so only the entered value is returned
disease = argument_dict["disease"].values.tolist()[0][0]
disease_omim = argument_dict["disease_omim"].values.tolist()[0][0]

# Extract embryo data into dataframe dropping any empty rows
embryo_data_df = argument_dict["embryo_data"].dropna(axis=0)
embryo_data_df.columns = [
    "biopsy_no",
    "embryo_id",
    "embryo_column_name",
]

exclusion = argument_dict["exclusion"]
female_partner_hosp_num = argument_dict["female_partner_hosp_num"]
flanking_region_size = argument_dict["flanking_region_size"]
gene = argument_dict["gene"]
gene_end = argument_dict["gene_end"]
gene_omim = argument_dict["gene_omim"]
gene_start = argument_dict["gene_start"]
input_file = argument_dict["input_file"]
maternal_mutation = argument_dict["maternal_mutation"]
mode_of_inheritance = argument_dict["mode_of_inheritance"]
multi_analysis = argument_dict["multi_analysis"]
(
    partner1_type,
    partner1_sex,
    partner1_forename,
    partner1_surname,
    partner1_dob,
    partner1_DNA_number,
    partner1_column_name,
) = argument_dict["partner1_details"].values.tolist()[0]
(
    partner2_type,
    partner2_sex,
    partner2_forename,
    partner2_surname,
    partner2_dob,
    partner2_DNA_number,
    partner2_column_name,
) = argument_dict["partner2_details"].values.tolist()[0]
paste_gene = argument_dict["paste_gene"]
paternal_mutation = argument_dict["paternal_mutation"]
pgd_worksheet = argument_dict["pgd_worksheet"]
pgd_worksheet_denovo = argument_dict["pgd_worksheet_denovo"]
pru = argument_dict["pru"]
reference = argument_dict["reference"]
ref_relationship = argument_dict["ref_relationship"]
ref_relationship_to_couple = argument_dict["ref_relationship_to_couple"]
ref_seq = argument_dict["ref_seq"]
ref_status = argument_dict["ref_status"]
template_version = argument_dict["template_version"]
