import csv
from io import StringIO
import openpyxl
import os

# Mapping the defined names to the old and new template names
text = """defined_name,old_template,new_template
biopsy_number,data_entry!$B$9,data_entry!$M$13
chromosome,data_entry!$B$31,data_entry!$B$30
consanguineous,data_entry!$H$15,data_entry!$H$15
de_novo,data_entry!$H$17,data_entry!$H$17
disease,data_entry!$F$22:$L$22,data_entry!$B$16
disease_omim,data_entry!$M$22:$N$22,data_entry!$C$26
embryo_data,data_entry!$Q$4:$T$28,data_entry!$J$23:$M$47
exclusion,data_entry!$M$17,data_entry!$M$17
female_partner_hosp_num,data_entry!$B$15,data_entry!$B$11
flanking_region_size,data_entry!$B$35,data_entry!$B$34
gene,data_entry!$B$27,data_entry!$B$26
gene_end,data_entry!$C$33,data_entry!$C$32
gene_omim,data_entry!$C$27,data_entry!$C$26
gene_start,data_entry!$B$33,data_entry!$B$32
input_file,data_entry!$B$7,data_entry!$M$11
maternal_mutation,data_entry!$B$20,data_entry!$B$19
mode_of_inheritance,data_entry!$B$3,data_entry!$B$3
multi_analysis,data_entry!$B$18,data_entry!$B$14
partner1_details,data_entry!$G$4:$M$4,data_entry!$G$4:$M$4
partner2_details,data_entry!$G$6:$M$6,data_entry!$G$6:$M$6
paste_gene,data_entry!$B$29,data_entry!$B$28
paternal_mutation,data_entry!$B$21,data_entry!$B$20
pgd_worksheet,data_entry!$B$5,data_entry!$B$5
pgd_worksheet_denovo,data_entry!$B$16,data_entry!$B$12
pru,data_entry!$B$14,data_entry!$B$10
reference,data_entry!$H$8:$M$8,data_entry!$H$8:$M$8
ref_relationship,data_entry!$H$12,data_entry!$H$12
ref_relationship_to_couple,data_entry!$H$13,data_entry!$H$13
ref_seq,data_entry!$B$22,data_entry!$B$21
ref_status,data_entry!$H$11,data_entry!$H$11
template_version,data_entry!$B$38,data_entry!$B$37"""

# Parse the CSV text
csv_data = csv.DictReader(StringIO(text))

# Process the CSV data
defined_names_mapping = {}
for row in csv_data:
    name = row["defined_name"]
    old_template = row["old_template"]
    new_template = row["new_template"]
    defined_names_mapping[name] = {
        "old_template": old_template,
        "new_template": new_template,
    }

# Load the new template workbook
new_template_path = "/home/graeme/Desktop/SNP_haplotyper/snp_haplotyper/excel_templates/basher_input_template_v2.0.0.xlsm"
new_wb = openpyxl.load_workbook(new_template_path)

# Loop through the defined names mapping
for name, new_template_ref in defined_names_mapping.items():
    # Remove the old defined name if it exists
    if name in new_wb.defined_names:
        del new_wb.defined_names[name]

    # Create a new defined name with the updated reference
    new_defined_name = openpyxl.workbook.defined_name.DefinedName(
        name=name, attr_text=new_template_ref["new_template"]
    )
    new_wb.defined_names.append(new_defined_name)


# Save the updated workbook
new_wb.save(new_template_path)


# Replace 'path/to/old/files' with the path to the directory containing old-format spreadsheets
old_files_directory = "/home/graeme/Desktop/SNP_haplotyper/test_data/template_test_data"

# List all files in the directory
old_files = [
    os.path.join(old_files_directory, f)
    for f in os.listdir(old_files_directory)
    if os.path.isfile(os.path.join(old_files_directory, f))
]

for old_file in old_files:
    # Load the old workbook and create a new workbook based on the updated template
    old_wb = openpyxl.load_workbook(old_file)
    new_wb = openpyxl.load_workbook(new_template_path, keep_vba=True)

    # Loop through the defined names mapping
    for name, refs in defined_names_mapping.items():
        # Get the old range and sheet
        old_range = list(old_wb.defined_names[name].destinations)[0]
        old_ws_name, old_cells_range = old_range
        old_ws = old_wb[old_ws_name]
        cells = old_ws[old_cells_range]

        # Get the new range and sheet
        new_range = list(new_wb.defined_names[name].destinations)[0]
        new_ws_name, new_cells_range = new_range
        new_ws = new_wb[new_ws_name]

        # Check if the range represents a single cell or multiple cells
        if name in [
            "disease",
            "disease_omim",
        ]:
            pass

        elif isinstance(cells, openpyxl.cell.cell.Cell):
            # It's a single cell, copy its value
            new_ws[new_cells_range].value = cells.value
        else:
            # It's a range of cells, copy the values
            first_cell_in_range = new_ws[new_cells_range][0][0]
            for row_idx, row in enumerate(cells):
                for col_idx, cell in enumerate(row):
                    new_ws.cell(
                        row=row_idx + first_cell_in_range.row,
                        column=col_idx + first_cell_in_range.column,
                    ).value = cell.value

    new_ws["$B$37"].value = "V2.0.0"

    # Replace 'path/to/new/files' with the path to the directory where you want to save the new files
    new_files_directory = (
        "/home/graeme/Desktop/SNP_haplotyper/test_data/template_test_data_v2"
    )
    new_file_name = os.path.splitext(os.path.basename(old_file))[0] + "_new.xlsx"
    new_file_path = os.path.join(new_files_directory, new_file_name)
    new_wb.save(new_file_path)
