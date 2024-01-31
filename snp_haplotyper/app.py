import logging
import os
import random
import tempfile
import time
import zipfile
from argparse import Namespace
from datetime import datetime
from io import BytesIO

import merge_array_files
import pdfkit
import snp_haplotype
from excel_parser import parse_excel_input
from flask import (
    Blueprint,
    Flask,
    Response,
    jsonify,
    render_template,
    request,
    send_file,
    session,
)
from flask_cors import CORS, cross_origin
from flask_session import Session
from flask_wtf import FlaskForm
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from werkzeug.utils import secure_filename
from wtforms import FileField, MultipleFileField, SubmitField, ValidationError

log_format = "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
logger = logging.getLogger("BASHer_logger")

# To override the default severity of logging
logger.setLevel("DEBUG")


# Create a new Flask web server instance
app = Flask(__name__)

# Define the folder where uploaded files will be stored. The folder location is retrieved from an environment variable
app.config["UPLOAD_FOLDER"] = os.environ["UPLOAD_FOLDER"]
app.config["SECRET_KEY"] = "catchmeifyoucan"
app.config["SESSION_TYPE"] = "filesystem"
app.config[
    "SESSION_PERMANENT"
] = True  # Set the session lifetime. If True, the session is permanent until the browser is closed.
app.config["SESSION_FILE_DIR"] = os.environ[
    "SESSION_FILE_DIR"
]  # Define the directory where session files will be stored.
app.config["UPLOAD_EXTENSIONS"] = [
    ".txt",
    ".csv",
    ".xlsm",
    ".xlsx",
]  # List of acceptable upload file extensions
app.config["MAX_CONTENT_LENGTH"] = (
    2 * 1024 * 1024
)  # Set the maximum size of uploaded files to 2MB
app.config["APPLICATION_ROOT"] = "/basher"  # Set the root URL of the application
app.config["WTF_CSRF_ENABLED"] = True

app.config.from_object(__name__)

IMG_VERSIONED = os.getenv(
    "IMG_VERSIONED", "N/A"
)  # Get the version of the application for display in

# Create a Blueprint object named "basher" that represents the "basher" component of the application. The URL prefix "/basher" is added to all routes defined in this blueprint.
basher_bp = Blueprint("basher", __name__, url_prefix="/basher")
Session(app)
CORS(
    app, supports_credentials=True
)  # Enable handling of cross-origin requests - required to run react components


class BasherForm(FlaskForm):
    sample_sheet = FileField("Upload Sample Sheet:")
    snp_array_files = MultipleFileField("Upload one or more SNP files:")
    submit = SubmitField("Run BASHer")


def call_excel_parser(sample_sheet, snp_array_file):
    # Get contents of parsed input sheet
    basher_input_namespace, error_dictionary, input_ok_flag = parse_excel_input(
        sample_sheet, snp_array_file
    )

    return [basher_input_namespace, error_dictionary, input_ok_flag]


def call_basher(basher_input_namespace):
    (
        mode_of_inheritance,
        sample_id,
        number_snps_imported,
        summary_snps_by_region,
        informative_snps_by_region,
        html_string,
        pdf_string,
    ) = snp_haplotype.main(basher_input_namespace)

    return sample_id, html_string, pdf_string


class ChangeForm(FlaskForm):
    sample_sheet = FileField(
        "Sample Sheet:",
        validators=[],
        id="sample_sheet",
        render_kw={"accept": ".xlsm, .xlsx"},
    )
    snp_array_files = MultipleFileField(
        "SNP Array Files:",
        validators=[],
        id="snp_array_files",
        render_kw={"accept": ".txt, .csv"},
    )
    submit = SubmitField("Run BASHer")

    def validate_sample_sheet(form, field):
        allowed_extensions = ["xlsm", "xlsx"]
        file = field.data
        if not file:
            raise ValidationError("No sample sheet provided.")
        if not file.filename:
            raise ValidationError("No sample sheet selected.")
        if not (
            "." in file.filename
            and file.filename.rsplit(".", 1)[1].lower() in allowed_extensions
        ):
            raise ValidationError(
                f"Invalid file type for sample sheet '{file.filename}'. Allowed types are: xlsm, xlsx"
            )
        # Check if the file is password protected
        try:
            # Convert SpooledTemporaryFile to BytesIO
            file_content = BytesIO(file.stream.read())
            # Reset the file position to the beginning
            file.stream.seek(0)
            workbook = load_workbook(file_content, read_only=True)
            sample_sheet_readable = True
        except (InvalidFileException, zipfile.BadZipFile):
            raise ValidationError(
                f"The sample sheet, {file.filename}, is password protected or corrupted."
            )
            sample_sheet_readable = False
            # Check if the file is password protected

        if sample_sheet_readable:
            # Convert SpooledTemporaryFile to BytesIO
            file_content = BytesIO(file.stream.read())
            # Reset the file position to the beginning
            file.stream.seek(0)
            workbook = load_workbook(file_content, read_only=True)
            # List of required defined names
            required_defined_names = [
                "biopsy_number",
                "chromosome",
                "consanguineous",
                "de_novo",
                "disease",
                "disease_omim",
                "embryo_data",
                "exclusion",
                "female_partner_hosp_num",
                "flanking_region_size",
                "gene",
                "gene_end",
                "gene_omim",
                "gene_start",
                "input_file",
                "maternal_mutation",
                "mode_of_inheritance",
                "multi_analysis",
                "partner1_details",
                "partner2_details",
                "paste_gene",
                "paternal_mutation",
                "pgd_worksheet",
                "pgd_worksheet_denovo",
                "pru",
                "reference",
                "ref_relationship",
                "ref_relationship_to_couple",
                "ref_seq",
                "ref_status",
                "template_version",
            ]
            # Get defined names
            defined_names = workbook.defined_names.definedName
            defined_names_list = [name.name for name in defined_names]
            # Check if required defined names exist and collect missing ones
            missing_names = [
                name
                for name in required_defined_names
                if name not in defined_names_list
            ]

            # Raise error if sample sheet is missing defined names
            if missing_names != []:
                raise ValidationError(
                    f'Defined names missing from the Excel file: {", ".join(missing_names)}. Please check the template file.'
                )

    def validate_snp_array_files(form, field):
        allowed_extensions = ["txt", "csv"]
        files = field.data
        if not files:
            raise ValidationError("No snp array file provided.")
        for file in files:
            if not file.filename:
                raise ValidationError("No snp array file selected.")
            if not (
                "." in file.filename
                and file.filename.rsplit(".", 1)[1].lower() in allowed_extensions
            ):
                raise ValidationError(
                    f"Invalid file type for SNP array file '{file.filename}'. Allowed types are: txt, csv"
                )


@basher_bp.route("/", methods=["GET", "POST"])
@cross_origin(supports_credentials=True)
def form(basher_state="initial"):
    """
    This function handles both GET and POST requests to the root ("/basher") of the "basher" blueprint.

    For GET requests, it just returns the main page with an empty form.

    For POST requests, which are initiated by the button in the HTML form, it processes the uploaded files, performs validation,
    initiates the required logging, manages the session state, calls other functions to parse Excel data and execute 'basher'
    operations, and finally renders the updated page, either with the result of the operation or with the validation errors.

    Parameters:
    basher_state (str): The initial state of the basher process. Defaults to "initial".

    Returns:
    A rendered template of the main page, including the form, the basher state, and any file errors.
    """
    SampleSheetUp = SampleSheetUpload()  # File Upload class - detailed below.
    SnpArrayUp = SnpArrayUpload()  # File Upload class - detailed below.
    chgForm = ChangeForm()
    chgDetail = dict()

    if request.method == "POST" and chgForm.validate_on_submit():
        if chgForm.errors:
            return render_template(
                "index.html",
                form=chgForm,
                file_errors=chgForm.errors,
            )

        # Use FileHandler() to log to a file

        session["timestr"] = datetime.now().strftime("%Y%m%d-%H%M%S")
        os.mkdir(os.path.join(app.config["UPLOAD_FOLDER"], session["timestr"]))
        file_handler = logging.FileHandler(f"/var/local/basher/logs/basher_error.log")
        formatter = logging.Formatter(log_format)
        file_handler.setFormatter(formatter)

        # Don't forget to add the file handler
        logger.addHandler(file_handler)

        # Saved with the session details to avoid duplicate files from other users/sessions
        # temporary files are saved so that they can be passed to the backend.
        sample_sheet = chgForm.sample_sheet.data
        input_sheet = SampleSheetUp.upload(sample_sheet)
        chgDetail["sample_sheet"] = input_sheet

        snp_array_files = chgForm.snp_array_files.data
        input_files = SnpArrayUp.upload(snp_array_files)
        # If multiple files are uploaded merge them into a single file, else just use the single file
        if len(input_files) > 1:
            # use removesuffix to remove .txt from the end of the file names in the list
            # and then join them together with a _ to make a new file name
            merged_file_name = (
                "_".join(
                    sorted(
                        [
                            os.path.splitext(os.path.basename(file_name))[0]
                            for file_name in input_files
                        ]
                    )
                )
                + "_merged.txt"
            )
            df = merge_array_files.main(input_files)
            df.to_csv(
                os.path.join(
                    app.config["UPLOAD_FOLDER"], session["timestr"], merged_file_name
                ),
                sep="\t",
            )
            input_file = merged_file_name
        else:
            input_file = input_files[0]

        chgDetail["snp_array_files"] = input_file
        basher_state = "started"  # This doesn't refer to the backend starting, rather that the data has been submit and controls the flow in the html

        # Get the file names of the uploaded files
        input_sheet_basename = os.path.basename(input_sheet)
        input_file_basename = os.path.basename(input_file)

        # Create the paths to the uploaded files
        input_sheet_tmp_path = os.path.join(
            app.config["UPLOAD_FOLDER"], session["timestr"], input_sheet_basename
        )

        input_file_tmp_path = os.path.join(
            app.config["UPLOAD_FOLDER"], session["timestr"], input_file_basename
        )

        basher_input_namespace, input_errors, input_ok_flag = call_excel_parser(
            input_sheet_tmp_path, input_file_tmp_path
        )

        if input_ok_flag == False:
            return render_template(
                "index.html",
                form=chgForm,
                basher_state="initial",
                errors=input_errors,  # Pass the errors to the template
                file_errors=chgForm.errors,
            )
        else:
            sample_id, html_report, pdf_report = call_basher(basher_input_namespace)

            session["report_name"] = f'{sample_id}_{session["timestr"]}'
            session["report_path"] = os.path.join(
                app.config["UPLOAD_FOLDER"],
                session["report_name"],
            )
            with open(
                f'{session["report_path"]}.html',
                "w",
            ) as f:
                f.write(html_report)
            logger.info(
                f"Saved HTML report for {sample_id} at {session['report_path']}.html"
            )

            # Convert HTML report to PDF
            pdfkit.from_string(
                pdf_report,
                f'{session["report_path"]}.pdf',
            )
            logger.info(f"Saved PDF report for {sample_id}")

            return render_template(
                "index.html",
                form=chgForm,
                basher_state=basher_state,
                sample_sheet_name=sample_sheet.filename,
                snp_array_file_names=", ".join([x.filename for x in snp_array_files]),
                report_name=f'{session["report_name"]}.html',
                file_errors=chgForm.errors,
            )

    return render_template(
        "index.html",
        form=chgForm,
        basher_state=basher_state,
        file_errors=chgForm.errors,
    )


class SampleSheetUpload:
    def upload(self, file):
        file_name = file.filename
        if file_name == "":
            return "NULL"

        else:
            secure_file_name = secure_filename(file_name)
            file.save(
                os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    session["timestr"],
                    secure_file_name,
                )
            )

            return str(
                os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    session["timestr"],
                    secure_file_name,
                )
            )


class SnpArrayUpload:
    def upload(self, files):
        file_names = []

        for file in files:
            file_name = file.filename
            if file_name == "":
                return "NULL"

            else:
                secure_file_name = secure_filename(file_name)
                file.save(
                    os.path.join(
                        app.config["UPLOAD_FOLDER"],
                        session["timestr"],
                        secure_file_name,
                    )
                )
                file_names.append(
                    str(
                        os.path.join(
                            app.config["UPLOAD_FOLDER"],
                            session["timestr"],
                            secure_file_name,
                        )
                    )
                )

        return file_names


@basher_bp.route("/download", methods=["GET", "POST"])
@cross_origin(supports_credentials=True)
def download():
    """
    This function handles both GET and POST requests to the "/download" route of the "basher" blueprint.

    Initiated by a button click in the HTML, it creates a zip file containing the HTML and PDF reports stored in the session.
    The function then removes the original HTML and PDF reports and sends the zip file to the client as a file download.

    Returns:
    A file download response containing the zip file with the HTML and PDF reports.
    """
    html_file_name = f'{session["report_name"]}.html'
    pdf_file_name = f'{session["report_name"]}.pdf'

    # Create a temporary directory
    with tempfile.TemporaryDirectory() as tempdir:
        # Create a zip file with the html and pdf reports
        zip_path = os.path.join(tempdir, f'{session["report_name"]}.zip')
        with zipfile.ZipFile(zip_path, "w") as zipObj:
            zipObj.write(f'{session["report_path"]}.html', html_file_name)
            zipObj.write(f'{session["report_path"]}.pdf', pdf_file_name)

        logger.info(f"Saved zipped reports to {zip_path}")

        # Delete the html and pdf reports
        os.remove(f'{session["report_path"]}.html')
        os.remove(f'{session["report_path"]}.pdf')

        logger.info(f"Attempting to download {zip_path}")

        return send_file(
            zip_path,
            download_name=f'{session["report_name"]}.zip',
            mimetype="application/zip",
            as_attachment=True,
        )


# Register the blueprint with your Flask application
app.register_blueprint(basher_bp)

if __name__ == "__main__":
    app.run(debug=True)
