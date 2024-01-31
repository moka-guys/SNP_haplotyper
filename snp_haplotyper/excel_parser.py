import argparse
import logging
import math
import os
import re
import subprocess
import sys
from pathlib import Path

import config as config
import pandas as pd
import snp_haplotype
from check_inputs import check_input
from EnumDataClasses import (
    Chromosome,
    FlankingRegions,
    InheritanceMode,
    Relationship,
    Sex,
    Status,
)
from helper_functions import get_clean_filename
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval


# Custom error handler which saves errors to a dictionary for feedback to user
class DictErrorHandler(logging.Handler):
    def __init__(self, error_dict):
        super().__init__()
        self.error_dict = error_dict

    def emit(self, record):
        if record.levelno == logging.ERROR:
            error_msg = self.format(record)
            if error_msg not in self.error_dict:
                self.error_dict[error_msg] = 0
            self.error_dict[error_msg] += 1


# Create an error dictionary
error_dict_parser = {}

# Initialize the custom error handler
dict_error_handler = DictErrorHandler(error_dict_parser)

# Configure the logger to use the custom handler
logger = logging.getLogger("BASHer_logger")
logger.setLevel(logging.ERROR)
logger.addHandler(dict_error_handler)


# Add the directory containing this script to the PYTHOPATH
sys.path.append(os.path.dirname(__file__))

# Import command line arguments (these can be automatically generated from the sample sheet using sample_sheet_reader.py)
parser = argparse.ArgumentParser(
    description="Parses meta data from provided excel file and runs SNP haplotyper"
)

# File input/output data
parser.add_argument(
    "-i",
    "--input_spreadsheet",
    type=str,
    help="Excel file containing SNP Array meta data",
)

parser.add_argument(
    "-s",
    "--snp_array_file",
    type=str,
    help="SNP Array text files to be processed",
    required=False,
    nargs="+",  # Allows for 0 or more arguments
    default=None,  # Default value if no arguments are provided
)

parser.add_argument(
    "--run_basher",
    action="store_true",
    help="If the flag is set then BASHer is automatically run with the parsed data",
)
parser.add_argument(
    "--parse_excel_only",
    dest="run_basher",
    action="store_false",
    help="If the flag is set then BASHer is not run, and the parsed data is returned",
)
parser.set_defaults(run_basher=True)


def load_workbook_range(range_string, worksheet):
    """
    Import rows in the range with text entered
    """
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in worksheet[range_string]:
        data_rows.append([cell.value for cell in row])

    df = pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

    # Convert float values to text as specified
    def convert_float_to_text(value):
        if isinstance(value, float):
            # Check for NaN
            if math.isnan(value):
                return (
                    None  # or return '' if you want to replace NaN with an empty string
                )
            # Remove the decimal point and append a zero
            return str(int(value))
        return value

    df = df.applymap(convert_float_to_text)

    return df


def parse_excel_input(input_spreadsheet, snp_array_file=None):
    """
    Imports the following defined cells/ranges from the provided excel file:
        biopsy_number
        chromosome
        consanguineous
        de_novo
        disease
        disease_omim
        embryo_data - range of cells containing embryo data:
            biopsy_no
            embryo_id
            embryo_sex
            embryo_column_name
        exclusion
        female_partner_hosp_num
        flanking_region_size
        gene
        gene_end
        gene_omim
        gene_start
        input_file
        maternal_mutation
        mode_of_inheritance
        multi_analysis
        partner1_details - range of cells containing partner1 details:
            partner1_type
            partner1_sex
            partner1_forename
            partner1_surname
            partner1_dob
            partner1_DNA_number
            partner1_column_name
        partner2_details - range of cells containing partner2 details:
            partner2_type
            partner2_sex
            partner2_forename
            partner2_surname
            partner2_dob
            partner2_DNA_number
            partner2_column_name
        paste_gene
        paternal_mutation
        pgd_worksheet
        pgd_worksheet_denovo
        pru
        reference  - range of cells containing reference details:
            reference_sex
            reference_forename
            reference_surname
            reference_dob
            reference_DNA_number
            reference_column_name
        ref_relationship
        ref_relationship_to_couple
        ref_seq
        ref_status
        template_version
    """
    wb = load_workbook(
        filename=input_spreadsheet,
        keep_vba=False,
        data_only=True,
        keep_links=True,
    )

    argument_dict = {}
    error_dict_parser = {}

    # Get list of defined ranges from provided excel sheet
    defined_ranges = wb.defined_names
    # TODO Add to logger when implemented
    #  print(defined_ranges)
    data_entry_sheet = wb["data_entry"]

    for dn in wb.defined_names.definedName:
        # TODO Add to logger when implemented
        # print(dn.name)
        # print(dn.attr_text)
        input_name = dn.name
        if input_name in [
            "embryo_data",
            "partner1_details",
            "partner2_details",
            "reference",
        ]:
            # Import excel ranges
            df = load_workbook_range(
                dn.attr_text.split("!")[1].replace("$", ""), data_entry_sheet
            )
            argument_dict[input_name] = df.dropna(how="all")  # Remove empty rows
        else:
            # Process cell locations in the format data_entry!$B$31 or data_entry!$F$22:$L$22 (merged cells)
            cell_location = dn.attr_text.split(":")[0].split("!")[1].replace("$", "")
            cell_value = str(data_entry_sheet[cell_location].value).strip()
            argument_dict[input_name] = cell_value

    biopsy_number = argument_dict["biopsy_number"]
    chr = Chromosome["CHR_" + str(argument_dict["chromosome"]).upper()]
    consanguineous = argument_dict["consanguineous"]
    de_novo = argument_dict["de_novo"]
    disease = argument_dict["disease"]
    disease_omim = argument_dict["disease_omim"]
    exclusion = argument_dict["exclusion"]
    female_partner_hosp_num = argument_dict["female_partner_hosp_num"]
    # flanking_region_size = argument_dict["flanking_region_size"]
    flanking_region_size = FlankingRegions.FLANK_2MB
    gene_symbol = argument_dict["gene"]
    gene_end = int(argument_dict["gene_end"])
    gene_omim = argument_dict["gene_omim"]
    gene_start = int(argument_dict["gene_start"])
    input_file = argument_dict["input_file"]
    maternal_mutation = argument_dict["maternal_mutation"]
    mode_of_inheritance = InheritanceMode(argument_dict["mode_of_inheritance"].lower())
    multi_analysis = argument_dict["multi_analysis"]
    paste_gene = argument_dict[
        "paste_gene"
    ]  # genomic range in format chr3:100000-200000 used to populate other fields
    paternal_mutation = argument_dict["paternal_mutation"]
    pgd_worksheet = argument_dict["pgd_worksheet"]
    pgd_worksheet_denovo = argument_dict["pgd_worksheet_denovo"]
    pru = argument_dict["pru"]
    reference = argument_dict["reference"]
    ref_relationship = argument_dict["ref_relationship"].lower()
    ref_relationship_to_couple = (
        None
        if argument_dict["ref_relationship_to_couple"] is None
        else argument_dict["ref_relationship_to_couple"].lower()
    )
    ref_seq = argument_dict["ref_seq"]
    ref_status_str = argument_dict["ref_status"].lower()

    # Convert the status string to Status enum
    try:
        ref_status = Status(ref_status_str)
    except ValueError:
        print(f"Invalid status: {ref_status_str}")

    template_version = argument_dict["template_version"]

    embryo_data_df = argument_dict["embryo_data"]
    column_names = ["biopsy_no", "embryo_id", "embryo_sex", "embryo_column_name"]

    # Ensure all values are strings
    embryo_data_df = embryo_data_df.astype(str)

    # Check that the embryo data sheet has the correct columns
    if len(embryo_data_df.columns) == len(column_names):
        embryo_data_df.columns = column_names

        # Ensure that "embryo_sex" is in lowercase
        if "embryo_sex" in embryo_data_df.columns:
            embryo_data_df["embryo_sex"] = embryo_data_df["embryo_sex"].str.lower()
        else:
            logging.error('Key "embryo_sex" is missing from the DataFrame')

        # If data has been passed in for the embryo data sheet, check that the selected biopsy number is in the sheet
        if (
            not embryo_data_df.empty
            and biopsy_number not in embryo_data_df["biopsy_no"].to_list()
        ):
            logger.error(
                f"Error: Biopsy number {biopsy_number} is not in the embryo data sheet."
            )
        # Filter embryo data to only include embryos from the current biopsy
        filtered_embryo_data_df = embryo_data_df[
            embryo_data_df["biopsy_no"] == biopsy_number
        ]

        # Check if there are any embryos in the current biopsy
        if filtered_embryo_data_df.empty:
            trio_only = True
        else:
            trio_only = False

    else:
        logger.error(
            f"Warning: Expected {len(column_names)} columns, {column_names} but got {len(embryo_data_df.columns)} columns."
        )

    (
        partner1_type,
        partner1_sex_str,
        partner1_forename,
        partner1_surname,
        partner1_dob,
        partner1_DNA_number,
        partner1_column_name,
    ) = argument_dict["partner1_details"].values.tolist()[0]

    # Convert the sex string to Sex enum
    try:
        partner1_sex = Sex(partner1_sex_str.lower())
    except ValueError:
        print(f"Invalid sex: {partner1_sex_str}")

    (
        partner2_type,
        partner2_sex_str,
        partner2_forename,
        partner2_surname,
        partner2_dob,
        partner2_DNA_number,
        partner2_column_name,
    ) = argument_dict["partner2_details"].values.tolist()[0]

    # Convert the sex string to Sex enum
    try:
        partner2_sex = Sex(partner2_sex_str.lower())
    except ValueError:
        print(f"Invalid sex: {partner2_sex_str}")

    (
        reference_sex_str,
        reference_forename,
        reference_surname,
        reference_dob,
        reference_DNA_number,
        reference_column_name,
    ) = argument_dict["reference"].values.tolist()[0]

    # Check if reference_sex_str is not None
    if reference_sex_str is None:
        reference_sex = Sex("unknown")
    else:
        try:
            reference_sex = Sex(reference_sex_str.lower())
        except ValueError:
            print(f"Invalid sex: {reference_sex_str}")

    if partner1_sex == Sex.MALE and partner2_sex == Sex.FEMALE:
        male_partner_status = Status(partner1_type.split("_")[0])
        male_partner_col = partner1_column_name
        female_partner_status = Status(partner2_type.split("_")[0])
        female_partner_col = partner2_column_name
    elif partner1_sex == Sex.FEMALE and partner2_sex == Sex.MALE:
        male_partner_status = Status(partner2_type.split("_")[0])
        male_partner_col = partner2_column_name
        female_partner_status = Status(partner1_type.split("_")[0])
        female_partner_col = partner1_column_name

    # TODO rationalise output prefix - input_file or input_spreadsheet - move cleanup of names into test function?
    output_prefix = (
        os.path.splitext(os.path.basename(input_spreadsheet))[0]
        .replace("excel_test_Autosomal_Dominant_", "")
        .replace("excel_test_Autosomal_Recessive_", "")
        .replace("excel_test_X_linked_", "")
    )

    # if mode_of_inheritance == InheritanceMode.AUTOSOMAL_DOMINANT:
    #     female_partner_status = female_partner_status
    #     male_partner_status = convert_to_status_enum(male_partner_status.split("_")[0])
    # elif mode_of_inheritance == InheritanceMode.AUTOSOMAL_RECESSIVE:
    #     female_partner_status = (
    #         convert_to_status_enum("carrier")
    #         if female_partner_status == "carrier_partner"
    #         else female_partner_status
    #     )
    #     male_partner_status = (
    #         convert_to_status_enum("carrier")
    #         if male_partner_status == "carrier_partner"
    #         else male_partner_status
    #     )
    # elif mode_of_inheritance == InheritanceMode.X_LINKED:
    #     female_partner_status = (
    #         convert_to_status_enum("carrier")
    #         if female_partner_status == "carrier_female_partner"
    #         else female_partner_status
    #     )
    #     male_partner_status = (
    #         convert_to_status_enum("unaffected")
    #         if male_partner_status == "unaffected_male_partner"
    #         else male_partner_status
    #     )

    lookup_dict = {
        "son": "child",
        "daughter": "child",
        "prenatal": "child",
        "embryo": "child",
        "mother": "grandparent",
        "father": "grandparent",
        "child": "child",
    }
    if ref_relationship.lower() in [
        "son",
        "daughter",
        "prenatal",
        "embryo",
        "mother",
        "father",
        "child",
    ]:
        ref_relationship = Relationship(lookup_dict[ref_relationship.lower()])

    # Export data as dictionary to be used in other functions & testing
    excel_import = {}
    excel_import["biopsy_number"] = biopsy_number
    excel_import["chr"] = chr
    excel_import["consanguineous"] = consanguineous
    excel_import["de_novo"] = de_novo
    excel_import["disease"] = disease
    excel_import["disease_omim"] = disease_omim
    excel_import["biopsy_no"] = biopsy_number
    excel_import["embryo_data"] = embryo_data_df
    excel_import["exclusion"] = exclusion
    excel_import["female_partner_hosp_num"] = female_partner_hosp_num
    # excel_import["flanking_region_size"] = flanking_region_size
    excel_import["flanking_region_size"] = FlankingRegions.FLANK_2MB
    excel_import["gene"] = gene_symbol
    excel_import["gene_end"] = gene_end
    excel_import["gene_omim"] = gene_omim
    excel_import["gene_start"] = gene_start
    excel_import["input_file"] = input_file
    excel_import["maternal_mutation"] = maternal_mutation
    excel_import["mode_of_inheritance"] = mode_of_inheritance
    excel_import["multi_analysis"] = multi_analysis
    excel_import["partner1_type"] = partner1_type
    excel_import["partner1_sex"] = partner1_sex
    excel_import["partner1_forename"] = partner1_forename
    excel_import["partner1_surname"] = partner1_surname
    excel_import["partner1_dob"] = partner1_dob
    excel_import["partner1_DNA_number"] = partner1_DNA_number
    excel_import["partner1_column_name"] = partner1_column_name
    excel_import["partner2_type"] = partner2_type
    excel_import["partner2_sex"] = partner2_sex
    excel_import["partner2_forename"] = partner2_forename
    excel_import["partner2_surname"] = partner2_surname
    excel_import["partner2_dob"] = partner2_dob
    excel_import["partner2_DNA_number"] = partner2_DNA_number
    excel_import["partner2_column_name"] = partner2_column_name
    excel_import["male_partner_status"] = male_partner_status
    excel_import["male_partner_col"] = male_partner_col
    excel_import["female_partner_status"] = female_partner_status
    excel_import["female_partner_col"] = female_partner_col
    excel_import["paste_gene"] = paste_gene
    excel_import["paternal_mutation"] = paternal_mutation
    excel_import["pgd_worksheet"] = pgd_worksheet
    excel_import["pgd_worksheet_denovo"] = pgd_worksheet_denovo
    excel_import["pru"] = pru
    excel_import["reference"] = reference
    excel_import["reference_sex"] = reference_sex
    excel_import["reference_forename"] = reference_forename
    excel_import["reference_surname"] = reference_surname
    excel_import["reference_dob"] = reference_dob
    excel_import["reference_DNA_number"] = reference_DNA_number
    excel_import["reference_column_name"] = reference_column_name
    excel_import["ref_relationship"] = ref_relationship
    excel_import["ref_relationship_to_couple"] = ref_relationship_to_couple
    excel_import["ref_seq"] = ref_seq
    excel_import["ref_status"] = ref_status
    excel_import["template_version"] = template_version

    # The user can specify the SNP array text files in both the provided template and the via the command line
    # If a SNP array text file is specified in both the template and the command line the two files must be the same
    # If a SNP array text file is specified in the template but not the command line, the file specified in the template will be used

    # Check whether a SNP array text file has been specified on the commandline, if they have then check
    # it against that provided in the template. If they are different, raise an error
    if snp_array_file is not None:
        # Get the base name of the snp_array_file
        snp_array_file_basename = os.path.basename(snp_array_file)

        # Get the list of input files from the excel_import
        input_files = [file.strip() for file in excel_import["input_file"].split(",")]

        # Remove the file extensions, sort the filenames, and concatenate the file names
        merged_name = (
            "_".join(
                sorted(
                    [
                        os.path.splitext(os.path.basename(file_name))[0]
                        for file_name in input_files
                    ]
                )
            )
            + "_merged.txt"
        )

        # Check if the snp_array_file is a merged file
        is_merged_file = snp_array_file_basename == merged_name

        # Check if the snp_array_file is not one of the input files and it's not a merged file
        if not is_merged_file and snp_array_file_basename not in [
            get_clean_filename(file_name) for file_name in input_files
        ]:
            logger.error(
                f"The SNP array text file specified on the command line, {snp_array_file_basename} is different to that specified in the template, {input_files} which is converted to {merged_name}."
            )

        input_filepath = snp_array_file

    else:
        # Check whether environment variable is set
        if "UPLOAD_FOLDER" in os.environ:
            # If docker-compose has set the environment variable, use the path specified in the environment variable
            input_filepath = os.path.join(
                os.environ["UPLOAD_FOLDER"], os.path.basename(input_file)
            )
        else:
            # If we are running outside of docker-compose, use the path specified in the config file
            input_filepath = input_file

    # Create an argparse and populate it with the required arguments for passing to snp_haplotyper
    args = argparse.Namespace()
    args.embryo_ids = embryo_data_df.embryo_id.to_list()
    args.embryo_sex = embryo_data_df.embryo_sex.to_list()
    args.mode_of_inheritance = mode_of_inheritance
    args.input_file = input_filepath
    args.output_folder = config.output_folder
    args.output_prefix = output_prefix
    args.mode_of_inheritance = mode_of_inheritance
    args.male_partner = male_partner_col
    args.male_partner_status = male_partner_status
    args.female_partner = female_partner_col
    args.female_partner_status = female_partner_status
    args.reference = reference_column_name
    args.reference_status = ref_status
    args.reference_sex = reference_sex
    args.reference_relationship = ref_relationship
    args.gene_symbol = gene_symbol
    args.gene_start = gene_start
    args.gene_end = gene_end
    args.chr = chr
    args.flanking_region_size = flanking_region_size
    args.consanguineous = True if consanguineous == "yes" else False

    # If analysis is being done for embryos add that data as well
    if trio_only == False:
        args.trio_only = False
        args.embryo_ids = filtered_embryo_data_df.embryo_column_name.to_list()
        args.embryo_sex = filtered_embryo_data_df.embryo_sex.to_list()
    else:
        args.trio_only = True

    # Add header info to cmd string
    args.header_info = (
        f"PRU={pru};Hospital No={female_partner_hosp_num};Biopsy No={biopsy_number}"
    )

    # Use check_inputs.py module to ensure the input data is valid
    error_dictionary, input_ok_flag = check_input(args, input_filepath)

    if error_dict_parser != {}:
        input_ok_flag = False
        error_dictionary.update(error_dict_parser)

    if input_ok_flag == False:
        logger.error("Input data is not valid, exiting - check log file for details")
        # sys.exit(1)
    else:
        logger.info("Input data from excel template has passed data validation checks")
    return args, error_dictionary, input_ok_flag


def main(excel_parser_args):
    # If the user has specified the run_basher flag, then parse the excel input and run snp_haplotyper
    if excel_parser_args.run_basher:
        if excel_parser_args.snp_array_file is None:
            excel_import = parse_excel_input(excel_parser_args.input_spreadsheet)
        else:
            excel_import = parse_excel_input(
                excel_parser_args.input_spreadsheet, excel_parser_args.snp_array_file
            )
        snp_haplotype.main(excel_import)
    # If the user has not specified the run_basher flag, then just parse the excel input
    else:
        excel_import = parse_excel_input(excel_parser_args.input_spreadsheet)
        return excel_import


if __name__ == "__main__":
    excel_parser_args = parser.parse_args()
    main(excel_parser_args)
